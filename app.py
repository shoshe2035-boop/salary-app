import streamlit as st
from datetime import date, timedelta
import pandas as pd
import io
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------
# إعدادات الصفحة والأنماط
# ---------------------------------------------------------
st.set_page_config(page_title="نظام الفروقات الجماعي", layout="wide")

# تحسين الأنماط لظهور النصوص في الجداول بشكل داكن
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Cairo:wght@400;700&display=swap');
    html, body, .stApp { 
        direction: rtl !important; 
        text-align: right !important; 
        background-color: #f8f9fa !important; 
        font-family: 'Cairo', sans-serif !important; 
        color: #212529 !important; 
    }
    [data-testid="stSidebar"] { 
        background-color: #1E1E2F !important; 
    }
    [data-testid="stSidebar"] * { 
        color: #FFFFFF !important; 
    }
    .stButton button { 
        background-color: #1E3A8A !important; 
        color: white !important; 
        border-radius: 8px; 
        border: none; 
    }
    .stButton button:hover { 
        background-color: #0F2B5C !important; 
    }
    div[data-baseweb="input"] input, .stNumberInput input, .stDateInput input, .stSelectbox div {
        background-color: white !important; 
        color: black !important; 
        border: 1px solid #ced4da !important; 
        direction: rtl !important;
    }
    label p { 
        color: #1E3A8A !important; 
        font-weight: 600; 
    }
    /* تحسين جداول البيانات */
    .stDataFrame, .stTable {
        color: black !important;
    }
    .stDataFrame table, .stTable table {
        color: black !important;
        background-color: white !important;
    }
    .stDataFrame td, .stTable td {
        color: black !important;
    }
    .printable-report { 
        background-color: white !important; 
        padding: 30px; 
        border: 2px solid #1E3A8A; 
        margin-bottom: 20px; 
        page-break-after: always; 
    }
    .report-table { 
        width: 100%; 
        border-collapse: collapse; 
        margin-top: 10px; 
    }
    .report-table th { 
        background-color: #1E3A8A !important; 
        color: white !important; 
        border: 1px solid #dee2e6; 
        padding: 8px; 
    }
    .report-table td { 
        border: 1px solid #dee2e6; 
        padding: 8px; 
        color: black !important; 
    }
    @media print {
        .no-print { display: none !important; }
        .printable-report { border: 2px solid #000 !important; box-shadow: none; page-break-after: always; }
    }
</style>
""", unsafe_allow_html=True)

# ---------------------------------------------------------
# إدارة حالة الجلسة
# ---------------------------------------------------------
if 'all_employees' not in st.session_state:
    st.session_state.all_employees = []
if 'current_actions' not in st.session_state:
    st.session_state.current_actions = []
if 'form_counter' not in st.session_state:
    st.session_state.form_counter = 0

def reset_form():
    st.session_state.current_actions = []
    st.session_state.form_counter += 1

# ---------------------------------------------------------
# دوال الحساب
# ---------------------------------------------------------
def calculate_employee(emp):
    rates = {"بكالوريوس": 0.45, "دبلوم": 0.55, "ماجستير": 0.75, "دكتوراه": 1.0, "اعدادية": 0.25, "متوسطة": 0.15}
    rate = rates.get(emp['degree'], 0)
    rows_html = ""
    total_nominal = 0
    cumulative_diff = 0
    prev_salary = emp['base']
    prev_year = None

    actions = emp['actions']
    for i, act in enumerate(actions):
        diff = act['salary'] - prev_salary
        is_new_year = (prev_year is not None and act['date'].year > prev_year)

        if emp['mode'] == "تراكم مستمر (مضاعفة دائماً)" or is_new_year:
            effective_diff = diff + cumulative_diff
            note = "تراكمي" if cumulative_diff > 0 else "بداية"
        else:
            effective_diff = diff
            note = "نفس السنة"

        cumulative_diff += diff

        next_date = actions[i+1]['date'] if i < len(actions)-1 else emp['end']
        start = act['date']
        if start.day >= 25:
            start = (start.replace(day=1) + timedelta(days=32)).replace(day=1)
        months = (next_date.year - start.year) * 12 + (next_date.month - start.month)
        if months < 0:
            months = 0

        if months > 0:
            row_nom = effective_diff * months
            total_nominal += row_nom
            rows_html += f"<tr><td>{act['type']}</td><td>{act['order']}</td><td>{months}</td><td>{effective_diff:,}</td><td>{row_nom:,}</td><td>{note}</td></tr>"

        prev_salary = act['salary']
        prev_year = act['date'].year

    net = total_nominal * rate
    return rows_html, total_nominal, net

# ---------------------------------------------------------
# دالة إنشاء ملف Excel مع تنسيق جاهز للطباعة
# ---------------------------------------------------------
def generate_excel():
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        workbook = writer.book
        # تنسيق الرأس
        header_font = Font(bold=True, size=12)
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for emp in st.session_state.all_employees:
            # تنظيف اسم الورقة من الأحرف غير المسموح بها
            sheet_name = emp['name'].replace('/', '_').replace('\\', '_').replace('*', '_').replace('?', '_').replace(':', '_')[:30]
            
            # بيانات الموظف (صف واحد)
            df_emp = pd.DataFrame([{
                'الموظف': emp['name'],
                'المؤهل': emp['degree'],
                'الراتب الأساسي': emp['base'],
                'تاريخ النهاية': emp['end'],
                'طريقة الاحتساب': emp['mode']
            }])
            df_emp.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0)

            # تنسيق صف البيانات
            worksheet = writer.sheets[sheet_name]
            for col in range(1, 6):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border

            # حركات الموظف (جدول)
            actions_data = []
            for act in emp['actions']:
                actions_data.append({
                    'نوع الحركة': act['type'],
                    'رقم الأمر': act['order'],
                    'الراتب الجديد': act['salary'],
                    'التاريخ': act['date']
                })
            df_actions = pd.DataFrame(actions_data)
            df_actions.to_excel(writer, sheet_name=sheet_name, index=False, startrow=3)

            # تنسيق جدول الحركات
            for col_idx, col_name in enumerate(df_actions.columns, 1):
                cell = worksheet.cell(row=4, column=col_idx)
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border
                # ضبط عرض العمود
                column_letter = get_column_letter(col_idx)
                worksheet.column_dimensions[column_letter].width = max(len(col_name) + 5, 15)
            
            for row_idx in range(5, 5 + len(df_actions)):
                for col_idx in range(1, 5):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')

            # نتائج الحساب
            _, total_nom, total_net = calculate_employee(emp)
            df_result = pd.DataFrame([{
                'الاسمي الكلي': total_nom,
                'الصافي المستحق': total_net
            }])
            df_result.to_excel(writer, sheet_name=sheet_name, index=False, startrow=8)

            # تنسيق النتائج
            for col_idx, col_name in enumerate(df_result.columns, 1):
                cell = worksheet.cell(row=9, column=col_idx)
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border
            for row_idx in range(10, 10 + len(df_result)):
                for col_idx in range(1, 3):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    # تنسيق الأرقام
                    cell.number_format = '#,##0'

            # إضافة خط فاصل وإجمالي
            worksheet.cell(row=12, column=1, value="------------------------").font = Font(bold=True)

    output.seek(0)
    return output

# ---------------------------------------------------------
# الشريط الجانبي
# ---------------------------------------------------------
with st.sidebar:
    st.header("⚙️ الإعدادات")
    calc_mode = st.radio("طريقة الاحتساب:", ["تراكم مستمر (مضاعفة دائماً)", "مضاعفة في السنة الجديدة فقط"])
    st.divider()
    st.subheader(f"عدد الموظفين: {len(st.session_state.all_employees)}")
    
    if st.button("🖨️ طباعة التقارير", use_container_width=True):
        st.markdown("<script>window.print();</script>", unsafe_allow_html=True)
    
    if st.session_state.all_employees:
        try:
            excel_data = generate_excel()
            st.download_button(
                label="📥 تصدير إلى Excel (مرتب)",
                data=excel_data,
                file_name=f"تقرير_الفروقات_{date.today()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        except Exception as e:
            st.error(f"حدث خطأ أثناء إنشاء ملف Excel: {e}")
    
    if st.button("🗑️ تصفير الكل", use_container_width=True):
        st.session_state.all_employees = []
        reset_form()
        st.rerun()
    st.divider()
    st.caption("المبرمج: مصطفى حسن")

# ---------------------------------------------------------
# العنوان الرئيسي
# ---------------------------------------------------------
st.markdown('<h1 style="text-align:center; color:#1E3A8A;">📊 نظام احتساب الفروقات الوظيفية</h1>', unsafe_allow_html=True)

# ---------------------------------------------------------
# نموذج إدخال موظف جديد
# ---------------------------------------------------------
form_key = f"form_{st.session_state.form_counter}"

col_new, col_spacer = st.columns([1, 5])
with col_new:
    if st.button("➕ موظف جديد", type="primary", use_container_width=True):
        reset_form()
        st.rerun()

with st.container():
    st.markdown("### ➕ إضافة موظف جديد")

    col1, col2 = st.columns(2)
    with col1:
        name = st.text_input("اسم الموظف (اختياري)", key=f"name_{form_key}")
        base_sal = st.number_input("الراتب الأساسي (بالآلاف)", min_value=0, value=0, key=f"base_{form_key}") * 1000
    with col2:
        degree = st.selectbox("التحصيل العلمي", ["بكالوريوس", "دبلوم", "ماجستير", "دكتوراه", "اعدادية", "متوسطة"], key=f"deg_{form_key}")
        end_date = st.date_input("تاريخ نهاية الاحتساب", value=date.today(), key=f"end_{form_key}")

    st.markdown("#### حركات الموظف (علاوات/ترفيعات)")
    a1, a2, a3, a4 = st.columns([2, 2, 2, 2])
    with a1:
        act_type = st.selectbox("نوع الحركة", ["علاوة سنوية", "ترفيع وظيفي"], key=f"type_{form_key}")
    with a2:
        act_order = st.text_input("رقم الأمر", key=f"order_{form_key}")
    with a3:
        act_salary = st.number_input("الراتب الجديد (آلاف)", min_value=0, value=0, key=f"sal_{form_key}") * 1000
    with a4:
        act_date = st.date_input("التاريخ", key=f"date_{form_key}")

    col_btn1, col_btn2 = st.columns([1, 1])
    with col_btn1:
        if st.button("➕ إضافة الحركة", use_container_width=True):
            if act_salary > 0 and act_date:
                st.session_state.current_actions.append({
                    "type": act_type,
                    "order": act_order if act_order.strip() else "بدون أمر",
                    "salary": act_salary,
                    "date": act_date
                })
                st.session_state.current_actions.sort(key=lambda x: x['date'])
                st.rerun()
            else:
                st.warning("الراتب والتاريخ مطلوبان")

    with col_btn2:
        if st.button("💾 حفظ الموظف", use_container_width=True, type="primary"):
            if st.session_state.current_actions:
                emp_name = name.strip() if name.strip() else "موظف بدون اسم"
                st.session_state.all_employees.append({
                    "name": emp_name,
                    "degree": degree,
                    "base": base_sal,
                    "end": end_date,
                    "actions": st.session_state.current_actions.copy(),
                    "mode": calc_mode
                })
                reset_form()
                st.rerun()
            else:
                st.error("يجب إضافة حركة واحدة على الأقل قبل الحفظ")

# عرض الحركات المضافة مؤقتاً (باستخدام dataframe لضمان وضوح النص)
if st.session_state.current_actions:
    st.markdown("**الحركات المضافة لهذا الموظف:**")
    df_display = pd.DataFrame(st.session_state.current_actions)
    df_display.columns = ['النوع', 'رقم الأمر', 'الراتب', 'التاريخ']
    st.dataframe(df_display, use_container_width=True, hide_index=True)

# ---------------------------------------------------------
# عرض النتائج للموظفين المحفوظين
# ---------------------------------------------------------
if st.session_state.all_employees:
    st.divider()
    st.subheader("📋 الموظفين المضافة")

    for idx, emp in enumerate(st.session_state.all_employees):
        col_title, col_del = st.columns([6, 1])
        with col_title:
            st.markdown(f"**{idx+1}. {emp['name']}** ({emp['degree']})")
        with col_del:
            if st.button("🗑️ حذف", key=f"del_{idx}"):
                st.session_state.all_employees.pop(idx)
                st.rerun()

    # توليد التقارير القابلة للطباعة
    for emp in st.session_state.all_employees:
        rows, total_nom, total_net = calculate_employee(emp)
        st.markdown(f"""
        <div class="printable-report">
            <div style="display: flex; justify-content: space-between; border-bottom: 2px solid #1E3A8A; padding-bottom: 10px;">
                <h3>مديرية تربية الديوانية</h3>
                <h3>كشف فروقات رواتب</h3>
            </div>
            <div style="display: flex; justify-content: space-around; margin: 10px 0; font-weight: bold;">
                <span>الموظف: {emp['name']}</span>
                <span>المؤهل: {emp['degree']}</span>
                <span>تاريخ الغلق: {emp['end']}</span>
            </div>
            <table class="report-table">
                <thead><tr><th>الحركة</th><th>رقم الأمر</th><th>أشهر</th><th>الفرق</th><th>الاسمي</th><th>ملاحظة</th></tr></thead>
                <tbody>{rows}</tbody>
            </table>
            <div style="margin-top: 20px; display: flex; justify-content: space-between;">
                <span style="font-weight: bold;">مجموع الفروقات الاسمية: {total_nom:,} د.ع</span>
                <span style="font-weight: bold; color: #1E3A8A;">الصافي المستحق: {total_net:,.0f} د.ع</span>
            </div>
            <div style="margin-top: 40px; display: flex; justify-content: space-around;">
                <div>التدقيق: ________</div>
                <div>المدير المالي: ________</div>
                <div>المحاسب: ________</div>
            </div>
        </div>
        """, unsafe_allow_html=True)
